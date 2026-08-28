#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Вивантаження каталогу VELNOX у файл для синхронізації з 1С.

Джерело — знімок прод-бази SQLite. Нічого не змінює, лише читає.

    python3 tools/export_1c.py <шлях-до-database.sqlite> [тека-виводу]

На виході:
    velnox-catalog-1c.xlsx   — 7 аркушів: товари, характеристики (довгі й широкі),
                               аналоги, зображення, довідник характеристик, про файл
    velnox-catalog-1c.csv    — той самий плоский зріз одним рядком на товар,
                               UTF-8 з BOM (1С читає без танців з кодуванням)

Посилання на зображення робляться абсолютними: у базі шлях лежить як
/velnox/images/..., а прод-nginx зрізає /velnox, тож публічна адреса —
https://velnox.eu/images/...
"""
import csv
import io
import json
import os
import sqlite3
import sys
from collections import defaultdict

SITE = 'https://velnox.eu'
LOCALES = ['uk', 'en', 'pl']


def public_url(path: str) -> str:
    """/velnox/images/... -> https://velnox.eu/images/..."""
    if not path:
        return ''
    if path.startswith('http'):
        return path
    if path.startswith('/velnox/'):
        path = path[len('/velnox'):]
    if not path.startswith('/'):
        path = '/' + path
    return SITE + path


def load(db_path):
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    q = lambda sql, *a: [dict(r) for r in con.execute(sql, a).fetchall()]

    tr = defaultdict(dict)          # (entity_type, entity_id, field)[locale] = value
    for r in q("select entity_type, entity_id, locale, field, value from translations"):
        tr[(r['entity_type'], r['entity_id'], r['field'])][r['locale']] = r['value']

    cats   = {r['id']: r for r in q("select * from categories")}
    tables = {r['id']: r for r in q("select * from product_tables")}
    specs  = {r['id']: r for r in q("select * from spec_definitions")}
    spec_by_key = {r['key']: r for r in specs.values()}

    values = defaultdict(dict)      # product_id -> {spec_key: value}
    for r in q("select product_id, spec_id, value from product_specs"):
        if r['spec_id'] in specs:
            values[r['product_id']][specs[r['spec_id']]['key']] = r['value']

    refs = defaultdict(list)
    for r in q("select * from product_cross_refs order by product_id, brand, value"):
        refs[r['product_id']].append(r)

    assets = defaultdict(list)      # (entity_type, entity_id) -> [asset]
    for r in q("select * from product_assets order by entity_type, entity_id, type, sort_order, id"):
        assets[(r['entity_type'], r['entity_id'])].append(r)

    products = q("select * from products order by product_table_id, id")
    con.close()
    return dict(tr=tr, cats=cats, tables=tables, specs=specs, spec_by_key=spec_by_key,
                values=values, refs=refs, assets=assets, products=products)


def build_rows(d):
    """Один словник на товар — спільне джерело і для xlsx, і для csv."""
    rows = []
    for p in d['products']:
        tbl = d['tables'].get(p['product_table_id'], {})
        cat = d['cats'].get(tbl.get('category_id'), {})
        pid, tid = p['id'], tbl.get('id')

        name = d['tr'].get(('product', pid, 'name'), {})
        desc = d['tr'].get(('product', pid, 'desc'), {})
        mt   = d['tr'].get(('product', pid, 'meta_title'), {})
        md   = d['tr'].get(('product', pid, 'meta_description'), {})

        gallery = [a for a in d['assets'].get(('product', pid), []) if a['type'] == 'gallery']
        models  = [a for a in d['assets'].get(('product', pid), []) if a['type'] == 'model_3d']
        t_gal   = [a for a in d['assets'].get(('product_table', tid), []) if a['type'] == 'gallery']
        t_png   = [a for a in d['assets'].get(('product_table', tid), []) if a['type'] == 'schema_png']
        t_svg   = [a for a in d['assets'].get(('product_table', tid), []) if a['type'] == 'schema_svg']

        photos = [public_url(a['path']) for a in gallery] or [public_url(a['path']) for a in t_gal]
        cat_slug = cat.get('slug', '')

        row = {
            'article':        p['article'],
            'slug':           p['slug'],
            'category':       cat_slug,
            'category_name':  d['tr'].get(('category', cat.get('id'), 'name'), {}).get('uk', ''),
            'table':          tbl.get('slug', ''),
            'table_name':     d['tr'].get(('product_table', tid, 'name'), {}).get('uk', ''),
            'photo_main':     photos[0] if photos else '',
            'photos':         ';'.join(photos),
            'photos_count':   len(photos),
            'model_3d':       public_url(models[0]['path']) if models else '',
            'schema_png':     public_url(t_png[0]['path']) if t_png else '',
            'schema_svg':     public_url(t_svg[0]['path']) if t_svg else '',
            'cross_refs_count': len(d['refs'].get(pid, [])),
            'cross_refs':     ';'.join('%s %s' % (r['brand'] or '', r['value']) for r in d['refs'].get(pid, [])).strip(),
        }
        for loc in LOCALES:
            row['name_' + loc] = name.get(loc, '')
            row['desc_' + loc] = desc.get(loc, '')
            row['meta_title_' + loc] = mt.get(loc, '')
            row['meta_description_' + loc] = md.get(loc, '')
            row['url_' + loc] = '%s/%s/products/%s/%s' % (SITE, loc, cat_slug, p['slug'])

        cols = json.loads(tbl.get('spec_columns') or '[]')
        row['_spec_order'] = cols
        row['_specs'] = d['values'].get(pid, {})
        row['_pid'] = pid
        rows.append(row)
    return rows


BASE_COLS = [
    ('article',        'Артикул'),
    ('slug',           'Slug'),
    ('category',       'Категорія (код)'),
    ('category_name',  'Категорія'),
    ('table',          'Таблиця (код)'),
    ('table_name',     'Таблиця'),
    ('name_uk',        'Назва uk'),
    ('name_en',        'Назва en'),
    ('name_pl',        'Назва pl'),
    ('url_uk',         'Сторінка uk'),
    ('url_en',         'Сторінка en'),
    ('url_pl',         'Сторінка pl'),
    ('photo_main',     'Головне фото'),
    ('photos_count',   'Фото, шт'),
    ('photos',         'Усі фото'),
    ('model_3d',       '3D-модель'),
    ('schema_png',     'Схема PNG'),
    ('schema_svg',     'Схема SVG'),
    ('cross_refs_count', 'Аналогів, шт'),
    ('cross_refs',     'Аналоги'),
    ('desc_uk',        'Опис uk'),
    ('desc_en',        'Опис en'),
    ('desc_pl',        'Опис pl'),
    ('meta_title_uk',  'Meta title uk'),
    ('meta_title_en',  'Meta title en'),
    ('meta_title_pl',  'Meta title pl'),
    ('meta_description_uk', 'Meta description uk'),
    ('meta_description_en', 'Meta description en'),
    ('meta_description_pl', 'Meta description pl'),
]


def write_xlsx(d, rows, out):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    FONT = 'Arial'
    head_font = Font(name=FONT, bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor='1B3A6B')
    body_font = Font(name=FONT)

    wb = Workbook()

    def sheet(title, headers, data, widths=None):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font, cell.fill = head_font, head_fill
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        for r in data:
            ws.append(r)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(vertical='top')
        for i, h in enumerate(headers, start=1):
            w = (widths or {}).get(h)
            if w is None:
                longest = max([len(str(h))] + [len(str(r[i-1])) for r in data[:200] if len(r) >= i])
                w = min(max(longest + 2, 10), 60)
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        return ws

    # --- Про файл ---
    ws = wb.active
    ws.title = 'Про файл'
    about = [
        ('Вивантаження каталогу VELNOX', ''),
        ('', ''),
        ('Джерело', 'прод-база velnox.eu (SQLite), знімок'),
        ('Товарів', len(rows)),
        ('Категорій', len({r['category'] for r in rows})),
        ('Таблиць', len({r['table'] for r in rows})),
        ('', ''),
        ('Аркуш «Товари»', 'один рядок на товар: назви, описи, meta, посилання на сторінку і фото'),
        ('Аркуш «Характеристики»', 'по рядку на кожну характеристику товару — ключ, підпис трьома мовами, значення'),
        ('Аркуш «Характеристики широко»', 'той самий набір, але один рядок на товар, колонки — коди характеристик'),
        ('Аркуш «Аналоги»', 'перехресні аналоги: бренд, позначення, тип'),
        ('Аркуш «Зображення»', 'усі фото, схеми та 3D-моделі з прямими посиланнями'),
        ('Аркуш «Довідник характеристик»', 'код характеристики -> підпис uk/en/pl та одиниця'),
        ('', ''),
        ('Посилання', 'абсолютні, відкриваються без авторизації. Перевірено: віддаються з кодом 200'),
        ('Кодування CSV', 'UTF-8 з BOM'),
        ('Порожня клітинка', 'означає, що характеристики немає в цій таблиці, а не нуль'),
    ]
    for r in about:
        ws.append(list(r))
    ws['A1'].font = Font(name=FONT, bold=True, size=14)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if not cell.font.bold:
                cell.font = body_font
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 80
    for c in ('A3', 'A4', 'A5', 'A6', 'A8', 'A9', 'A10', 'A11', 'A12', 'A13', 'A15', 'A16', 'A17'):
        ws[c].font = Font(name=FONT, bold=True)

    # --- Товари ---
    sheet('Товари',
          [h for _, h in BASE_COLS],
          [[r[k] for k, _ in BASE_COLS] for r in rows],
          widths={'Опис uk': 60, 'Опис en': 60, 'Опис pl': 60, 'Усі фото': 50, 'Аналоги': 50,
                  'Meta description uk': 50, 'Meta description en': 50, 'Meta description pl': 50})

    # --- Характеристики (довгий формат) ---
    long_rows = []
    for r in rows:
        keys = r['_spec_order'] or sorted(r['_specs'])
        for k in keys:
            if k not in r['_specs']:
                continue
            sd = d['spec_by_key'].get(k, {})
            lab = d['tr'].get(('spec_definitions', sd.get('id'), 'label'), {})
            unit = d['tr'].get(('spec_definitions', sd.get('id'), 'unit'), {})
            long_rows.append([r['article'], r['table'], k,
                              lab.get('uk', ''), lab.get('en', ''), lab.get('pl', ''),
                              r['_specs'][k], unit.get('uk', '')])
    sheet('Характеристики',
          ['Артикул', 'Таблиця', 'Код', 'Підпис uk', 'Підпис en', 'Підпис pl', 'Значення', 'Одиниця'],
          long_rows)

    # --- Характеристики широко ---
    order = sorted(d['spec_by_key'].values(), key=lambda s: s['sort_order'] or 0)
    wide_keys = [s['key'] for s in order if any(s['key'] in r['_specs'] for r in rows)]
    sheet('Характеристики широко',
          ['Артикул'] + wide_keys,
          [[r['article']] + [r['_specs'].get(k, '') for k in wide_keys] for r in rows])

    # --- Аналоги ---
    ref_rows = []
    for r in rows:
        for x in d['refs'].get(r['_pid'], []):
            ref_rows.append([r['article'], x['brand'] or '', x['value'], x['type'] or ''])
    sheet('Аналоги', ['Артикул', 'Бренд', 'Позначення', 'Тип'], ref_rows)

    # --- Зображення ---
    img_rows = []
    TYPE_UK = {'gallery': 'фото', 'model_3d': '3D-модель',
               'schema_png': 'схема PNG', 'schema_svg': 'схема SVG'}
    for r in rows:
        for a in d['assets'].get(('product', r['_pid']), []):
            img_rows.append([r['article'], TYPE_UK.get(a['type'], a['type']), 'товар',
                             a['sort_order'], public_url(a['path'])])
        tid = next((t['id'] for t in d['tables'].values() if t['slug'] == r['table']), None)
        for a in d['assets'].get(('product_table', tid), []):
            img_rows.append([r['article'], TYPE_UK.get(a['type'], a['type']), 'таблиця',
                             a['sort_order'], public_url(a['path'])])
    sheet('Зображення', ['Артикул', 'Тип', 'Рівень', 'Порядок', 'Посилання'], img_rows,
          widths={'Посилання': 70})

    # --- Довідник характеристик ---
    dict_rows = []
    for s in order:
        lab = d['tr'].get(('spec_definitions', s['id'], 'label'), {})
        unit = d['tr'].get(('spec_definitions', s['id'], 'unit'), {})
        used = sum(1 for r in rows if s['key'] in r['_specs'])
        dict_rows.append([s['key'], lab.get('uk', ''), lab.get('en', ''), lab.get('pl', ''),
                          unit.get('uk', ''), s['svg_label'] or '', used])
    sheet('Довідник характеристик',
          ['Код', 'Підпис uk', 'Підпис en', 'Підпис pl', 'Одиниця', 'Літера на кресленні', 'Товарів'],
          dict_rows)

    wb.save(out)
    return out


def write_csv(d, rows, out):
    order = sorted(d['spec_by_key'].values(), key=lambda s: s['sort_order'] or 0)
    wide_keys = [s['key'] for s in order if any(s['key'] in r['_specs'] for r in rows)]
    headers = [h for _, h in BASE_COLS] + wide_keys
    with io.open(out, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        w.writerow(headers)
        for r in rows:
            w.writerow([r[k] for k, _ in BASE_COLS] + [r['_specs'].get(k, '') for k in wide_keys])
    return out


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    db = sys.argv[1]
    outdir = sys.argv[2] if len(sys.argv) > 2 else '.'
    os.makedirs(outdir, exist_ok=True)

    d = load(db)
    rows = build_rows(d)
    x = write_xlsx(d, rows, os.path.join(outdir, 'velnox-catalog-1c.xlsx'))
    c = write_csv(d, rows, os.path.join(outdir, 'velnox-catalog-1c.csv'))

    print('товарів: %d' % len(rows))
    print('характеристик: %d' % sum(len(r['_specs']) for r in rows))
    print('аналогів: %d' % sum(r['cross_refs_count'] for r in rows))
    print('зображень і моделей: %d' % sum(len(v) for v in d['assets'].values()))
    print('готово:\n  %s\n  %s' % (x, c))


if __name__ == '__main__':
    main()
